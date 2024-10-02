import openpyxl
import requests
import pandas as pd

def scan_file_hash(file_hash):
    """Scans a file hash using VirusTotal API and returns the result."""

    # Replace with your VirusTotal API key
    api_key = "WRITE YOUR API KEY HERE"

    url = f"https://www.virustotal.com/api/v3/files/{file_hash}"
    headers = {"x-apikey": api_key}
    response = requests.get(url, headers=headers)

    # Check for rate limit exceeded
    if response.status_code == 429:
        print("Rate limit exceeded. Waiting before retrying...")
        time.sleep(60)  # Wait for 1 minute and then retry

        # Retry the request after waiting
        response = requests.get(url, headers=headers)

    if response.status_code == 200:
        data = response.json()
        # Extract the file reputation from the json response from VT API
        file_reputation = data['data']['attributes']['reputation']
        # Extract the file score from the json response from VT API
        file_score = data['data']['attributes']['last_analysis_stats']['malicious']
        return file_reputation, file_score
    elif response.status_code == 404:
        # File hash not found
        print(f"File hash '{file_hash}' not found in VirusTotal.")
        file_reputation = None
        file_score = None
        return file_reputation, file_score
    else:
        print(f"Error: Received status code {response.status_code}")
        print(response.json())
        return None

if __name__ == "__main__":
    # Path to your Excel file
    file_path = r'DETERMINE THE FILE PATH HERE'
    
    # Read the Excel file
    #df = pd.read_excel(file_path, engine='openpyxl')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet['D1'] = 'File Reputation'
    sheet['E1'] = 'File Score'
    # Find the 'File Hashes' column by header name
    file_hashes_col = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == "File Hashes":  # First row is assumed to be the header row
            file_hashes_col = col[0].column
            break
    
    if file_hashes_col is None:
        print("Column 'File Hashes' not found.")
        
    
    # Iterate through the rows starting from the second row (ignoring the header)
    for row in sheet.iter_rows(min_row=2, min_col=file_hashes_col, max_col=file_hashes_col):

        # Rate limiting: max 4 requests per minute
        max_requests_per_minute = 4
        request_interval = 60 / max_requests_per_minute  # seconds per request

        file_hash = row[0].value  # Get the file hash value from the 'File Hashes' column
        if file_hash:
            # Call the other function with the file_hash value
            file_reputation, file_score = scan_file_hash(file_hash)
            if file_reputation is None and file_score is None: 
                # Write 'N/A' in the next right cell
                next_col = file_hashes_col + 1
                sheet.cell(row=row[0].row, column=next_col).value = "N/A"
                sheet.cell(row=row[0].row, column=next_col+1).value = "N/A"
            else:
                # Write the result in the next right cell
                next_col = file_hashes_col + 1
                sheet.cell(row=row[0].row, column=next_col).value = file_reputation
                sheet.cell(row=row[0].row, column=next_col+1).value = file_score

# Save the workbook
workbook.save(file_path)
print(f"Updated {file_path} successfully.")